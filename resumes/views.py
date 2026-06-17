import csv
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from .forms import CandidateStatusForm, ResumeUploadForm
from .models import Candidate, EmailLog
from .tasks import process_resume, send_candidate_status_email


@login_required
def resume_upload(request):
    form = ResumeUploadForm(request.POST or None, request.FILES or None, recruiter=request.user)
    if request.method == "POST" and form.is_valid():
        job = form.cleaned_data["job"]
        files = request.FILES.getlist("resumes")  # get all uploaded files

        if not files:
            messages.error(request, "Please upload at least one resume file.")
            return render(request, "resumes/resume_upload.html", {"form": form})

        count = 0
        for file in files:
            candidate = Candidate.objects.create(job=job, resume=file)
            process_resume.delay(candidate.id)
            count += 1

        messages.success(
            request,
            f"{count} resume(s) uploaded successfully. Parsing and scoring will run in the background.",
        )
        return redirect("resumes:candidate_list")

    return render(request, "resumes/resume_upload.html", {"form": form})


@login_required
def candidate_list(request):
    candidates = Candidate.objects.filter(job__recruiter=request.user).select_related("job")
    query = request.GET.get("q", "")
    status = request.GET.get("status", "")
    if query:
        candidates = candidates.filter(
            Q(name__icontains=query) | Q(email__icontains=query) | Q(skills__icontains=query)
        )
    if status:
        candidates = candidates.filter(status=status)
    return render(
        request,
        "resumes/candidate_list.html",
        {"candidates": candidates, "query": query, "status": status, "statuses": Candidate.Status.choices},
    )


@login_required
def candidate_detail(request, pk):
    candidate = get_object_or_404(Candidate.objects.select_related("job"), pk=pk, job__recruiter=request.user)
    form = CandidateStatusForm(instance=candidate)
    email_logs = candidate.email_logs.all()
    return render(
        request,
        "resumes/candidate_detail.html",
        {"candidate": candidate, "form": form, "email_logs": email_logs},
    )


@login_required
def candidate_status_update(request, pk):
    candidate = get_object_or_404(Candidate, pk=pk, job__recruiter=request.user)
    form = CandidateStatusForm(request.POST, instance=candidate)
    if form.is_valid():
        candidate = form.save()
        send_candidate_status_email.delay(candidate.id)
        messages.success(request, "Candidate status updated. Email task has been queued when applicable.")
    return redirect(candidate)


@login_required
def email_log_list(request):
    """Shows all email logs for candidates belonging to the logged-in recruiter."""
    logs = EmailLog.objects.filter(
        candidate__job__recruiter=request.user
    ).select_related("candidate", "candidate__job").order_by("-created_at")

    # Filter by status if provided
    status_filter = request.GET.get("status", "")
    if status_filter:
        logs = logs.filter(status=status_filter)

    return render(
        request,
        "resumes/email_log_list.html",
        {
            "logs": logs,
            "status_filter": status_filter,
            "statuses": EmailLog.Status.choices,
        },
    )
import csv
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone


@login_required
def export_candidates_csv(request):
    candidates = Candidate.objects.filter(
        job__recruiter=request.user
    ).select_related("job").order_by("-score")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="candidates_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    writer.writerow(["#", "Name", "Email", "Job", "Score (%)", "Experience (yrs)", "Skills", "Status", "Uploaded At"])

    for i, c in enumerate(candidates, 1):
        writer.writerow([
            i,
            c.name or "Unknown",
            c.email or "-",
            c.job.title,
            c.score,
            c.experience_years,
            c.skills or "-",
            c.get_status_display(),
            c.uploaded_at.strftime("%d %b %Y %H:%M"),
        ])

    return response


@login_required
def export_candidates_excel(request):
    candidates = Candidate.objects.filter(
        job__recruiter=request.user
    ).select_related("job").order_by("-score")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Styles
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="4F46E5")
    center_align = Alignment(horizontal="center", vertical="center")
    border_side  = Side(style="thin", color="E2E8F0")
    cell_border  = Border(bottom=Border(bottom=border_side).bottom)

    headers = ["#", "Name", "Email", "Job", "Score (%)", "Experience (yrs)", "Skills", "Status", "Uploaded At"]
    col_widths = [5, 25, 30, 25, 12, 16, 40, 14, 20]

    # Write headers
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30

    # Status fill colors
    status_colors = {
        "shortlisted": "D1FAE5",
        "rejected":    "FEE2E2",
        "review":      "EEF2FF",
    }

    # Write rows
    for i, c in enumerate(candidates, 1):
        row = i + 1
        values = [
            i,
            c.name or "Unknown",
            c.email or "-",
            c.job.title,
            float(c.score),
            c.experience_years,
            c.skills or "-",
            c.get_status_display(),
            c.uploaded_at.strftime("%d %b %Y %H:%M"),
        ]
        fill_color = status_colors.get(c.status, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            if col in [1, 5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Color status column
            if col == 8:
                cell.fill = row_fill
                cell.font = Font(bold=True)

        ws.row_dimensions[row].height = 22

    # Freeze header row
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="candidates_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response